#!/usr/bin/env python3
"""
writer_cogs.py – Tiefen-Debug (COGS 2) mit LLM-only Forecast
===========================================================
- Liest config/cogs2_accounts.csv für row→category
- Workbook A (data_only=True) zum Auslesen historischer Werte
- Workbook B (write-enabled) wird von main.py übergeben
- Schreibt LLM-Prognosen in Spalten t1–t3 und JSON in Spalte reason
- Logt in outputs/cogs2_debug.txt
"""
from __future__ import annotations
from pathlib import Path
from csv import DictReader
import re, json
from typing import List, Dict, Optional

from openpyxl import load_workbook
from loader import find_header_row, col_map
from explanations import explain

# ——————————————————————————————————————————————————————————————————— #
BASE     = Path(__file__).resolve().parent.parent.parent
MAP_CSV  = BASE / "config"   / "cogs2_accounts.csv"
SRC_XLSX = BASE / "data"     / "UnternehmensplanungExcel.xlsx"
LOG_FILE = BASE / "outputs"  / "cogs2_debug.txt"
SHEET    = "COGS (2)"

PLACEH   = {None, "", "-", "???", "."}
LOG: List[str] = []
def log(msg: str) -> None:
    LOG.append(msg)

_rx = re.compile(r"[^\d,.\-]")
def safe_float(v) -> Optional[float]:
    if v in PLACEH:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = _rx.sub("", v).replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None

def write_cogs_forecast(wb) -> None:
    # 1) Mapping einlesen
    if not MAP_CSV.exists():
        print("❌ Mapping CSV fehlt – bitte discover_accounts.py ausführen.")
        return

    cfg: Dict[int, str] = {
        int(r["row"]): r["category"].strip().lower()
        for r in DictReader(MAP_CSV.open(encoding="utf-8"))
    }
    log(f"Mapping geladen: {len(cfg)} Einträge")

    # 2) Workbooks und Sheets
    data_ws = load_workbook(SRC_XLSX, data_only=True)[SHEET]
    ws      = wb[SHEET]

    # 3) Header-Zeile finden
    header = find_header_row(ws)
    if header is None:
        log("ERROR: Header nicht gefunden.")
        _write_log()
        return
    log(f"Header-Zeile: {header}")

    # 4) Spalten-Mapping mit col_map
    cols = col_map(ws, header)  # {'t-2':3, 't-1':4, 't0':5, 't1':6, 't2':7, 't3':8}
    COL_T0     = cols["t0"]
    COL_FC     = {k: cols[k] for k in ("t1", "t2", "t3")}
    COL_REASON = max(COL_FC.values()) + 1
    log(f"Spalten-Mapping: {cols}, reason in {COL_REASON}")

    # 5) Konto-Spalte ermitteln (erste Nicht-Leer unter Header, links von t-2)
    acc_col = next(
        c
        for c in range(1, cols["t-2"])
        if any(
            isinstance(ws.cell(r, c).value, str)
            and ws.cell(r, c).value.strip()
            for r in range(header + 1, header + 8)
        )
    )
    log(f"Kontospalte erkannt: {acc_col}")

    # 6) Durch alle Forecast-Zeilen iterieren
    writes = 0
    for row, cat in sorted(cfg.items()):
        if cat != "forecast":
            log(f"Skip row {row} (category={cat})")
            continue

        # historische Werte aus data_ws
        t2 = safe_float(data_ws.cell(row, cols["t-2"]).value)
        t1 = safe_float(data_ws.cell(row, cols["t-1"]).value)
        t0 = safe_float(data_ws.cell(row, COL_T0).value)
        log(f"ROW {row} | t-2={t2} | t-1={t1} | t0={t0}")
        if t0 is None:
            log(f"  -> t0 fehlt, skip row {row}")
            continue

        # LLM anwerfen
        raw = explain(ws.cell(row, acc_col).value, [t2 or 0, t1 or 0, t0], [])
        log(f"  RAW_LLM row {row}: {raw!r}")

        # JSON parsen und schreiben
        try:
            obj = json.loads(raw)
            for key, col in COL_FC.items():
                v = obj.get(key)
                if isinstance(v, (int, float)):
                    ws.cell(row, col, value=round(v, 2))
                    writes += 1
            ws.cell(row, COL_REASON, value=obj.get("reason", ""))
        except Exception as e:
            log(f"  ERROR parsing JSON for row {row}: {e}")

    log(f"TOTAL writes={writes}")
    _write_log()


def _write_log() -> None:
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    LOG_FILE.write_text("\n".join(LOG), encoding="utf-8")
