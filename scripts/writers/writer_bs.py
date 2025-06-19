"""
writer_bs.py – Tiefen-Debug (BS 2) mit Doppel-Strategie und LLM-only Forecast
===========================================================================

- Liest config/bs2_accounts.csv für row→category
- Workbook A (data_only=True) zum Auslesen historischer Werte
- Workbook B (write-enabled) wird von main.py übergeben
- Schreibt LLM-Prognosen in Spalten F–H (t1–t3) und JSON in Spalte I
- Logt die Schritte in outputs/bs2_debug.txt
"""

from __future__ import annotations
from pathlib import Path
from csv import DictReader
import re
import json
import warnings
from typing import List

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from loader import find_header_row
from forecast import cagr, project
from explanations import explain

# --------------------------------------------------------------------------- #
#  Pfade & Konstanten                                                         #
# --------------------------------------------------------------------------- #
BASE      = Path(__file__).resolve().parent.parent.parent
MAP_CSV   = BASE / "config"  / "bs2_accounts.csv"
SRC_XLSX  = BASE / "data"    / "UnternehmensplanungExcel.xlsx"
LOG_FILE  = BASE / "outputs" / "bs2_debug.txt"

SHEET     = "BS (2)"
COL_T0    = column_index_from_string("E")
COL_FC    = {
    "t1": column_index_from_string("F"),
    "t2": column_index_from_string("G"),
    "t3": column_index_from_string("H"),
}
COL_REASON = COL_FC["t3"] + 1

PLACEH    = {None, "", "-", "???", "."}

LOG: List[str] = []
def log(msg: str) -> None:
    LOG.append(msg)

# --------------------------------------------------------------------------- #
#  safe_float                                                                 #
# --------------------------------------------------------------------------- #
_rx = re.compile(r"[^\d,.\-]")
def safe_float(v) -> float | None:
    if v in PLACEH:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = _rx.sub("", v).replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None

# --------------------------------------------------------------------------- #
#  Konto-Spalte erkennen                                                      #
# --------------------------------------------------------------------------- #
def detect_acc_col(ws, header_row: int, max_col: int = 15) -> int | None:
    for col in range(1, max_col + 1):
        if any(
            isinstance(ws.cell(r, col).value, str)
            and ws.cell(r, col).value.strip()
            for r in range(header_row + 1, header_row + 8)
        ):
            return col
    return None

# --------------------------------------------------------------------------- #
#  Hauptfunktion                                                              #
# --------------------------------------------------------------------------- #
def write_bs_forecast(wb) -> None:
    """
    Forecast für Sheet BS (2) im übergebenen Workbook wb schreiben.
    """
    if not MAP_CSV.exists():
        print("❌ Mapping CSV fehlt – bitte discover_accounts.py ausführen.")
        return

    # 1) Mapping einlesen
    cfg = {
        int(row["row"]): row["category"].strip().lower()
        for row in DictReader(MAP_CSV.open(encoding="utf-8"))
    }
    log(f"Mapping geladen: {len(cfg)} Einträge")

    # 2) Historische Werte aus Quelldatei
    data_wb = load_workbook(SRC_XLSX, data_only=True)
    data_ws = data_wb[SHEET]

    # 3) Ziel-Sheet im übergebenen Workbook
    ws = wb[SHEET]

    # 4) Header-Zeile finden
    header_row = find_header_row(ws)
    if header_row is None:
        log("ERROR: Header-Zeile mit 't0' nicht gefunden.")
        _write_log()
        return
    log(f"Header-Zeile: {header_row}")

    # 5) Konto-Spalte finden
    acc_col = detect_acc_col(ws, header_row)
    if acc_col is None:
        log("ERROR: Kontospalte nicht erkannt.")
        _write_log()
        return
    log(f"Kontospalte erkannt: {acc_col}")

    # 6) Forecast pro Zeile
    writes = 0
    for r, category in cfg.items():
        if category != "forecast":
            log(f"Skip row {r} (category={category})")
            continue

        # a) historische Werte
        t2 = safe_float(data_ws.cell(r, COL_T0 - 2).value)
        t1 = safe_float(data_ws.cell(r, COL_T0 - 1).value)
        t0 = safe_float(data_ws.cell(r, COL_T0    ).value)
        log(f"ROW {r} | t-2={t2} | t-1={t1} | t0={t0}")

        if t2 is None or t0 is None:
            log(f"  -> BAD DATA in row {r}")
            continue

        # b) LLM-Aufruf
        raw_json = explain(ws.cell(r, acc_col).value, [t2, t1, t0], [])
        log(f"  RAW_LLM row {r}: {raw_json}")

        # c) JSON parsen & schreiben
        try:
            obj = json.loads(raw_json)
            for key in ("t1","t2","t3"):
                val = obj.get(key)
                if isinstance(val, (int, float)):
                    ws.cell(r, COL_FC[key], value=round(val, 2))
                    writes += 1
            ws.cell(r, COL_REASON, value=str(obj.get("reason","")))
        except Exception as e:
            log(f"  ERROR parsing JSON in row {r}: {e!s}")

    log(f"TOTAL writes={writes}")
    _write_log()

# --------------------------------------------------------------------------- #
#  Logfile schreiben                                                           #
# --------------------------------------------------------------------------- #
def _write_log() -> None:
    LOG_FILE.parent.mkdir(exist_ok=True, parents=True)
    LOG_FILE.write_text("\n".join(LOG), encoding="utf-8")
