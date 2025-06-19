"""
writer_rev_sbe.py – Forecast REV_sbE (2) via LLM
"""
from __future__ import annotations
from pathlib import Path
from csv import DictReader
import re, json
from typing import List
from openpyxl import load_workbook
from loader import find_header_row, col_map
from explanations import explain

BASE       = Path(__file__).resolve().parent.parent.parent
MAP_CSV    = BASE / "config"  / "revsbe2_accounts.csv"
SRC_XLSX   = BASE / "data"    / "UnternehmensplanungExcel.xlsx"
LOG_FILE   = BASE / "outputs" / "rev_sbe2_debug.txt"
SHEET      = "REV_sbE (2)"

PLACEH     = {None,"","-","—",".","…"}

LOG: List[str] = []
log = LOG.append

_rx = re.compile(r"[^\d,.\-]")
def safe_float(v):           # wie gewohnt …
    if v in PLACEH: return None
    if isinstance(v,(int,float)): return float(v)
    if isinstance(v,str):
        s=_rx.sub("",v).replace(".","").replace(",",".")
        try: return float(s)
        except: return None
    return None

def write_rev_sbe_forecast(wb) -> None:
    if not MAP_CSV.exists():
        print("❌ Mapping CSV fehlt – discover_accounts.py laufen lassen."); return

    cfg = {int(r["row"]): r["category"].strip().lower()
           for r in DictReader(MAP_CSV.open(encoding="utf-8"))}
    log(f"Mapping: {len(cfg)} Einträge")

    data_ws = load_workbook(SRC_XLSX, data_only=True)[SHEET]
    ws      = wb[SHEET]

    header = find_header_row(ws)
    if header is None: log("Header nicht gefunden"); _flush(); return

    cols = col_map(ws, header)          # {'t-2':C, 't-1':D, 't0':E, 't1':F, ...}
    COL_T0 = cols["t0"]
    COL_FC = {k: cols[k] for k in ("t1","t2","t3")}
    COL_REASON = max(COL_FC.values())+1

    # Konto-Spalte
    acc_col = next(c for c in range(1,16)
                   if any(isinstance(ws.cell(r,c).value,str) and ws.cell(r,c).value.strip()
                          for r in range(header+1, header+8)))

    writes = 0
    for r,cat in cfg.items():
        if cat!="forecast": log(f"Skip {r} ({cat})"); continue

        t2 = safe_float(data_ws.cell(r, cols["t-2"]).value)
        t1 = safe_float(data_ws.cell(r, cols["t-1"]).value)
        t0 = safe_float(data_ws.cell(r, COL_T0).value)
        if t0 is None: log(f"Row {r}: t0 fehlt"); continue

        baseline = []  # hier könnten einfache Schätzungen rein
        raw = explain(ws.cell(r,acc_col).value, [t2 or 0,t1 or 0,t0], baseline)
        log(f"LLM row {r}: {raw}")

        try:
            obj = json.loads(raw)
            for k,c in COL_FC.items():
                v=obj.get(k); 
                if isinstance(v,(int,float)): ws.cell(r,c,value=round(v,2)); writes+=1
            ws.cell(r,COL_REASON,value=obj.get("reason",""))
        except Exception as e:
            log(f"JSON-Error row {r}: {e}")

    log(f"writes={writes}")
    _flush()

def _flush():
    LOG_FILE.parent.mkdir(exist_ok=True,parents=True)
    LOG_FILE.write_text("\n".join(LOG),encoding="utf-8")
