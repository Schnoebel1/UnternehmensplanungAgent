"""
writer_cfr.py – Tiefen-Debug (CFR 2) mit LLM-only Forecast
==========================================================
- Liest config/cfr2_accounts.csv für row→category
- Workbook A (data_only=True) zum Auslesen historischer Werte
- Workbook B (write-enabled) wird von main.py übergeben
- Schreibt LLM-Prognosen in Spalten t1–t3 und JSON in die erste freie Spalte
"""

from __future__ import annotations
from pathlib import Path
from csv import DictReader
import re, json
from typing import List

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from loader import find_header_row, col_map
from explanations import explain

# --------------------------------------------------------------------------- #
#  Pfade & Konstanten                                                         #
# --------------------------------------------------------------------------- #
BASE       = Path(__file__).resolve().parent.parent.parent
MAP_CSV    = BASE / "config"  / "cfr2_accounts.csv"
SRC_XLSX   = BASE / "data"    / "UnternehmensplanungExcel.xlsx"
LOG_FILE   = BASE / "outputs" / "cfr2_debug.txt"

SHEET      = "CFR (2)"
PLACEH     = {None, "", "-", "—", ".", "…"}

LOG: List[str] = []
def log(msg: str) -> None:
    LOG.append(msg)

# --------------------------------------------------------------------------- #
#  Helper                                                                     #
# --------------------------------------------------------------------------- #
_rx = re.compile(r"[^\d,.\-]")
def safe_float(v) -> float | None:
    if v in PLACEH:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = _rx.sub("", v).replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None

def detect_acc_col(ws, header_row: int, max_col: int = 15) -> int | None:
    """erste Spalte unterhalb Header-Zeile mit nicht-leeren Strings"""
    for col in range(1, max_col + 1):
        if any(
            isinstance(ws.cell(r, col).value, str)
            and ws.cell(r, col).value.strip()
            for r in range(header_row + 1, header_row + 8)
        ):
            return col
    return None

# --------------------------------------------------------------------------- #
#  Hauptfunktion                                                              #
# --------------------------------------------------------------------------- #
def write_cfr_forecast(wb) -> None:
    if not MAP_CSV.exists():
        print("❌ Mapping CSV fehlt – bitte discover_accounts.py ausführen.")
        return

    cfg = {int(r["row"]): r["category"].strip().lower()
           for r in DictReader(MAP_CSV.open(encoding="utf-8"))}
    log(f"Mapping geladen: {len(cfg)} Einträge")

    data_ws = load_workbook(SRC_XLSX, data_only=True)[SHEET]
    ws      = wb[SHEET]

    header_row = find_header_row(ws)
    if header_row is None:
        log("ERROR: Header-Zeile mit 't0' nicht gefunden."); _write_log(); return
    log(f"Header-Zeile: {header_row}")

    # dynamische Perioden-Spalten
    cols = col_map(ws, header_row)          # {'t-2':3, 't-1':4, 't0':5, 't1':6, ...}
    try:
        COL_T0     = cols["t0"]
        COL_FC     = {"t1": cols["t1"], "t2": cols["t2"], "t3": cols["t3"]}
    except KeyError as e:
        log(f"ERROR: Spalte {e.args[0]} nicht im Header gefunden."); _write_log(); return
    COL_REASON = max(COL_FC.values()) + 1   # erste freie Spalte rechts

    acc_col = detect_acc_col(ws, header_row)
    if acc_col is None:
        log("ERROR: Kontospalte nicht erkannt."); _write_log(); return
    log(f"Kontospalte erkannt: {acc_col}")

    # -----------------------------------------------------------------
    writes = 0
    for r, cat in cfg.items():
        if cat != "forecast":
            log(f"Skip row {r} (category={cat})"); continue

        t2 = safe_float(data_ws.cell(r, cols.get("t-2", 0)).value) if "t-2" in cols else None
        t1 = safe_float(data_ws.cell(r, cols.get("t-1", 0)).value) if "t-1" in cols else None
        t0 = safe_float(data_ws.cell(r, COL_T0).value)

        # t0 ist Pflicht – sonst keinen Forecast
        if t0 is None:
            log(f"ROW {r}: t0 fehlt → BAD DATA"); continue

        if t2 is None:
            log(f"ROW {r}: t-2 fehlt, setze 0.0 als Platzhalter.")
            t2 = 0.0
        if t1 is None:
            log(f"ROW {r}: t-1 fehlt, setze 0.0 als Platzhalter.")
            t1 = 0.0

        log(f"ROW {r} | t-2={t2} | t-1={t1} | t0={t0}")

        raw_json = explain(ws.cell(r, acc_col).value, [t2, t1, t0], [])
        log(f"  RAW_LLM row {r}: {raw_json}")

        try:
            obj = json.loads(raw_json)
            for key, col in COL_FC.items():
                val = obj.get(key)
                if isinstance(val, (int, float)):
                    ws.cell(r, col, value=round(val, 2))
                    writes += 1
            ws.cell(r, COL_REASON, value=str(obj.get("reason", "")))
        except Exception as e:
            log(f"  ERROR parsing/writing row {r}: {e!s}")

    log(f"TOTAL writes={writes}")
    _write_log()

# --------------------------------------------------------------------------- #
#  Logfile                                                                    #
# --------------------------------------------------------------------------- #
def _write_log() -> None:
    LOG_FILE.parent.mkdir(exist_ok=True, parents=True)
    LOG_FILE.write_text("\n".join(LOG), encoding="utf-8")
