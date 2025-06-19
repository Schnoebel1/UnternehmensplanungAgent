#!/usr/bin/env python3
"""
writer_staff.py – Tiefen-Debug (STAFF 2) mit LLM-only Forecast
=============================================================
- Liest config/staff2_accounts.csv für row→category
- Workbook A (data_only=True) zum Auslesen historischer Werte
- Workbook B (write-enabled) wird von main.py übergeben
- Schreibt LLM-Prognosen in Spalten t1–t3 und JSON in Spalte reason
- Logt in outputs/staff2_debug.txt UND druckt Debug-Infos auf die Konsole
"""
from __future__ import annotations
from pathlib import Path
from csv import DictReader
import re, json
from typing import List, Dict, Optional

from openpyxl import load_workbook
from loader import find_header_row
from explanations import explain

# ——————————————————————————————————————————————————————————————————— #
BASE       = Path(__file__).resolve().parent.parent.parent
MAP_CSV    = BASE / "config"  / "staff2_accounts.csv"
SRC_XLSX   = BASE / "data"    / "UnternehmensplanungExcel.xlsx"
LOG_FILE   = BASE / "outputs" / "staff2_debug.txt"
SHEET      = "STAFF (2)"

PLACEH: set = {None, "", "-", "???", "."}
LOG: List[str] = []

def log(msg: str) -> None:
    LOG.append(msg)
    print(msg)  # echo to console

_rx = re.compile(r"[^\d,.\-]")
def safe_float(v) -> Optional[float]:
    if v in PLACEH:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = _rx.sub("", v).replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None

def write_staff_forecast(wb) -> None:
    # 1) Mapping einlesen
    if not MAP_CSV.exists():
        print("❌ Mapping CSV fehlt – bitte discover_accounts.py ausführen.")
        return
    cfg: Dict[int, str] = {
        int(r["row"]): r["category"].strip().lower()
        for r in DictReader(MAP_CSV.open(encoding="utf-8"))
    }
    log(f"Mapping geladen: {len(cfg)} Einträge")

    # 2) Sheets öffnen
    data_ws = load_workbook(SRC_XLSX, data_only=True)[SHEET]
    ws      = wb[SHEET]

    # 3) DEBUG: Vorschau der ersten 5 Zeilen
    print("\n--- SHEET PREVIEW (erste 5 Zeilen) ---")
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        print(f"{i:2d}:", row[:10])
        if i >= 5:
            break
    print("--- end preview ---\n")

    # 4) Header-Zeile finden (nur "Gesamt 12/t0")
    header = find_header_row(ws, ["Gesamt 12/t0"])
    if header is None:
        log("ERROR: Header nicht gefunden.")
        _write_log()
        return
    log(f"Header-Zeile: {header}")

    # 5) Spalten-Mapping per manueller Suche
    def find_col(sub: str) -> int:
        for c in range(1, ws.max_column+1):
            val = ws.cell(header, c).value
            if isinstance(val, str) and sub.lower() in val.lower():
                return c
        raise KeyError(f"Spalte '{sub}' nicht gefunden in Header")

    COL_T0     = find_col("Gesamt 12/t0")
    COL_FC     = {
        "t1": find_col("t1"),
        "t2": find_col("t2"),
        "t3": find_col("t3"),
    }
    COL_REASON = max(COL_FC.values()) + 1
    log(f"Spalten gefunden: t0={COL_T0}, t1={COL_FC['t1']}, t2={COL_FC['t2']}, t3={COL_FC['t3']}, reason={COL_REASON}")

    # 6) Konto-Spalte autodetect (erste Nicht-Leer links von t0)
    acc_col = next(
        c for c in range(1, COL_T0)
        if any(
            isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.strip()
            for r in range(header + 1, header + 8)
        )
    )
    log(f"Kontospalte erkannt: {acc_col}")

    # 7) Forecast-Loop
    writes = 0
    for row, cat in sorted(cfg.items()):
        if cat != "forecast":
            log(f"Skip row {row} (category={cat})")
            continue

        t2 = safe_float(data_ws.cell(row, COL_T0-2).value)
        t1 = safe_float(data_ws.cell(row, COL_T0-1).value)
        t0 = safe_float(data_ws.cell(row, COL_T0).value)
        log(f"ROW {row} | t-2={t2} | t-1={t1} | t0={t0}")
        if t0 is None:
            log(f"  -> t0 fehlt, skip row {row}")
            continue

        acc_text = ws.cell(row, acc_col).value
        log(f"  Konto-Text: {acc_text!r}")
        raw = explain(acc_text, [t2 or 0, t1 or 0, t0], [])
        log(f"  RAW_LLM row {row}: {raw!r}")

        try:
            obj = json.loads(raw)
            for key, col in COL_FC.items():
                v = obj.get(key)
                if isinstance(v, (int, float)):
                    ws.cell(row, col, value=round(v, 2))
                    writes += 1
            reason = obj.get("reason", "")
            ws.cell(row, COL_REASON, value=reason)
            log(f"  → geschrieben t1–t3 + reason '{reason}'")
        except Exception as e:
            log(f"  ERROR parsing JSON row {row}: {e}")

    log(f"TOTAL writes = {writes}")
    _write_log()

def _write_log() -> None:
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    LOG_FILE.write_text("\n".join(LOG), encoding="utf-8")

# zum schnellen Testen
if __name__ == "__main__":
    from openpyxl import Workbook
    wb = Workbook()
    wb.create_sheet(SHEET)
    write_staff_forecast(wb)
