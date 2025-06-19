#!/usr/bin/env python3
"""
discover_accounts.py – Debug-Edition
====================================
• Erstellt <slug>_accounts.csv pro Sheet.
•   python discover_accounts.py                → normal
    python discover_accounts.py --debug REV   → debug nur dieses Sheet
"""
from __future__ import annotations
import unicodedata, yaml, re, sys
from difflib import SequenceMatcher
from pathlib import Path
from csv import DictWriter
from typing import Dict, List
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from loader import find_header_row

# --------------------------------------------------------------------------- #
BASE      = Path(__file__).resolve().parent.parent
SRC_XLSX  = BASE / "data"   / "UnternehmensplanungExcel.xlsx"
CFG_FILE  = BASE / "config" / "sheets.yml"
OUT_DIR   = BASE / "config"

# ---------- Normalisierung --------------------------------------------------
_ASCII_RE = re.compile(r"[a-z0-9]+")

def norm(s: str) -> str:
    """
    1) casefold(): macht Lowercase und wandelt 'ß' → 'ss'
    2) NFKD      : zerlegt ä/ö/ü in Basisbuchstaben + Kombinationszeichen
    3) ASCII     : entfernt die Kombinationszeichen
    4) regex     : behält nur a–z und 0–9
    """
    if not isinstance(s, str):
        return ""
    s1 = s.casefold()
    s2 = unicodedata.normalize("NFKD", s1)
    s3 = s2.encode("ascii", "ignore").decode("ascii")
    return "".join(_ASCII_RE.findall(s3))


# ---------- Fuzzy-Match -----------------------------------------------------
def similar(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


# --------------------------------------------------------------------------- #
def discover_sheet(name: str, spec: Dict, debug: bool=False) -> None:
    keys = [norm(a) for a in spec.get("forecast_accounts", [])]
    wb   = load_workbook(SRC_XLSX, read_only=True, data_only=True)
    if name not in wb.sheetnames:
        if debug: print(f"⚠️  Sheet '{name}' fehlt")
        return
    ws = wb[name]

    # 1) header finden
    header = find_header_row(ws, spec.get("header_aliases", ["t0"]))
    if not header:
        if debug: print(f"⚠️  '{name}': Header nicht gefunden")
        return

    # 2) account_column override?
    acc_col_spec = spec.get("account_column")
    if acc_col_spec:
        if isinstance(acc_col_spec, str):
            acc_col = column_index_from_string(acc_col_spec.upper())
        else:
            acc_col = int(acc_col_spec)
        if debug:
            print(f"→ Verwende CONFig-Spalte '{acc_col_spec}' (Index {acc_col}) für Konten-Texte")
    else:
        # 3) autodetect: erste Spalte mit Text unter dem Header
        acc_col = next(
            (
                c
                for c in range(1, 16)
                if any(
                    isinstance(ws.cell(r, c).value, str)
                    and ws.cell(r, c).value.strip()
                    for r in range(header + 1, header + 8)
                )
            ),
            None,
        )
        if debug:
            print(f"→ Autodetected acc_col = {acc_col}")

    if not acc_col:
        if debug: print(f"⚠️  '{name}': Konto-Spalte nicht gefunden")
        return

    rows: List[Dict[str,str]] = []
    txt_norms: List[str]    = []

    for r in range(header + 1, ws.max_row + 1):
        raw = ws.cell(r, acc_col).value
        if not isinstance(raw, str):
            continue
        txt = raw.strip()
        if not txt:
            continue

        # Debug: Rohtext und Norm
        n = norm(txt)
        if debug:
            print(f"\n[Row {r}] raw='{txt}' → norm='{n}'")

        # Debug: Similarities
        sims = [(key, similar(n, key)) for key in keys]
        if debug:
            for key, score in sims:
                mark = "✔" if score >= .90 else "✘"
                print(f"    {mark} sim('{n}','{key}') = {score:.1%}")

        cat = "forecast" if any(score >= .90 for _, score in sims) else "readonly"
        if debug:
            print(f"  → Kategorie: {cat.upper()}")

        txt_norms.append(n)
        rows.append({"row": r, "text": txt, "category": cat})

    # Gesamt-Debug-Report
    if debug:
        print(f"\n=== DEBUG-Report für Sheet '{name}' ===")
        for key in keys:
            best_score, best_norm = max(
                ((similar(key,n),n) for n in txt_norms), default=(0,"")
            )
            status = "OK" if best_score >= .90 else "MISS"
            print(f"[{status}] key='{key}'  best_norm='{best_norm}'  sim={best_score:.1%}")
        print("-"*60)

    # CSV-Ausgabe
    slug = norm(name)
    out  = OUT_DIR / f"{slug}_accounts.csv"
    OUT_DIR.mkdir(exist_ok=True)
    with out.open("w", newline="", encoding="utf-8") as f:
        w = DictWriter(f, ["row","text","category"])
        w.writeheader()
        w.writerows(rows)

    print(f"\n✅ '{name}': {len(rows)} Zeilen → {out.name}")


# --------------------------------------------------------------------------- #
def main():
    cfg      = yaml.safe_load(CFG_FILE.read_text(encoding="utf-8"))["sheets"]
    dbg_arg  = [a for a in sys.argv[1:] if not a.startswith("-")]
    debug_on = "--debug" in sys.argv

    if debug_on and dbg_arg:
        sheet_key = dbg_arg[0].lower()
        for name, spec in cfg.items():
            if sheet_key in name.lower():
                discover_sheet(name, spec, debug=True)
                break
        else:
            print("Sheet-Name nicht gefunden.")
    else:
        for name, spec in cfg.items():
            discover_sheet(name, spec, debug=debug_on)


if __name__ == "__main__":
    main()
