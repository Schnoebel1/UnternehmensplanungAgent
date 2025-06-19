"""
scan_bs_columns.py
------------------
Scant das Sheet 'BS (2)' und druckt die ersten 10 nicht-leeren Zellen jeder
Spalte (Zeilen 3-40). Damit siehst du sofort, welche Spalte die Kontobezeich-
nungen enthält (A=1, B=2, …).
"""

from pathlib import Path
from openpyxl import load_workbook

SRC = (Path(__file__).resolve().parent.parent / "data"
       / "UnternehmensplanungExcel.xlsx")
SHEET = "BS (2)"
MAX_ROWS = 40

wb = load_workbook(SRC, read_only=True, data_only=True)
ws = wb[SHEET]

print(f"--- Schneller Spalten-Scan ({SHEET}) ---")
for col in range(1, 8):         # A…G
    values = [ws.cell(r, col).value
              for r in range(3, MAX_ROWS)
              if isinstance(ws.cell(r, col).value, str) and ws.cell(r, col).value.strip()]
    if values:
        print(f"Spalte {col}:  {values[:10]}")
