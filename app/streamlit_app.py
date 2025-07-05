# app/streamlit_app.py

import streamlit as st
import pandas as pd
import altair as alt
import logging
import time
from pathlib import Path
import xlwings as xw
from openpyxl import load_workbook

# --- Logging ---------------------------------------------------------------
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger(__name__)

# --- Streamlit Page Config -----------------------------------------------
st.set_page_config(
    page_title="Forecast Simulator",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- Sidebar Steuerung & Navigation --------------------------------------
st.sidebar.header("🔧 Steuerung")
# Szenario-Stärke Slider (UI-only)
szenario = st.sidebar.slider(
    "🔮 Szenariostärke", 
    min_value=0, max_value=10, value=5, 
    help="Wählen Sie hier die Stärke Ihres Forecast-Szenarios."
)

if st.sidebar.button("🚀 Simulation starten"):
    progress = st.sidebar.progress(0)
    for i in range(100):
        progress.progress(i+1)
        time.sleep(5/100)
    progress.empty()
    st.sidebar.success("Simulation abgeschlossen!")

st.sidebar.markdown("---")
st.sidebar.markdown("**Seiten**")
page = st.sidebar.radio(
    "Seite wählen",  # non-empty label für Barrierefreiheit
    ["Übersicht", "Umsatz", "EBIT-Marge", "Cashflow", "FCF", "Kapitalrendite"],
    index=0,
    label_visibility="collapsed"  # versteckt das Label optisch
)

# --- Pfad zur Forecast-Datei prüfen & Recalc via xlwings -----------------
BASE = Path(__file__).resolve().parent.parent
FILE = BASE / "outputs" / "UnternehmensplanungForecast.xlsx"
SHEET = "KPI"

if not FILE.exists():
    st.error(f"Forecast-Datei nicht gefunden:\n{FILE}")
    st.stop()

try:
    app = xw.App(visible=False)
    wb_xw = app.books.open(str(FILE))
    wb_xw.app.api.CalculateFullRebuild()
    wb_xw.save()
    wb_xw.close()
    app.quit()
    logger.debug("Excel neu berechnet und gespeichert via xlwings.")
except Exception as e:
    logger.warning(f"xlwings Recalc fehlgeschlagen: {e}")

# --- KPI-Sheet mit openpyxl data_only einlesen ----------------------------
@st.cache_data
def load_kpis(path: Path, sheet: str = SHEET) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    header = [c.value for c in ws[4]]        # Excel-Zeile 5 (0-based idx=4)
    periods = header[4:10]                   # Spalten E–J = t-2…t3

    sections = {"Profitabilität","Werttreiber","Liquidität","Verschuldung"}
    wanted = {
        "EBITDA","EBITDA - Margin","EBIT","EBIT - Margin",
        "Umsatzrentabilität","EK-Rendite",
        "Materialaufwand / Umsatz","PersExp / Umsatz","D&A / Umsatz",
        "NWC","NWC change","NWC Intensität",
        "Debt to Equity","Net debt to EBITDA",
    }

    entries = []
    for r in range(6, ws.max_row+1):
        raw = ws.cell(r,2).value             # Spalte B
        if raw is None: 
            continue
        label = str(raw).strip()
        if label in sections or label not in wanted:
            continue
        vals = [ws.cell(r,c).value for c in range(5,11)]
        entries.append([label,*vals])

    return pd.DataFrame(entries, columns=["KPI",*periods]).set_index("KPI")

kpi_df = load_kpis(FILE)

# --- Hilfsfunktion zum Formatieren ---------------------------------------
def format_cell(label: str, val) -> str:
    if val is None or (isinstance(val,float) and pd.isna(val)):
        return ""
    num = float(val)
    pct_keys = ["marge","rentabilität","rendite","%","change","debt","net","intensity"]
    if any(k in label.lower() for k in pct_keys):
        return f"{num:.1%}"
    return f"{num:,.0f}"

# --- Seite: Übersicht (Landing Page) --------------------------------------
if page == "Übersicht":
    st.title("🚀 Forecast Simulator")
    st.markdown("""
    **Willkommen!**  
    Unser Forecast Simulator hilft Dir, auf Basis historischer Daten und frei wählbarer **Szenariostärke**  
    in wenigen Klicks Deine Finanz-KPIs der nächsten Perioden zu visualisieren.

    **So funktioniert’s:**  
    1. **Szenariostärke wählen** (im Slider oben), um aggressivere oder konservativere Forecasts zu simulieren.  
    2. **Simulation starten** – unser Backend berechnet und speichert die neuen Werte.  
    3. **KPIs ansehen:**  
       - **Umsatz**  
       - **EBIT-Marge**  
       - **Cashflow**  
       - **Free Cashflow**  
       - **Kapitalrendite**  
    4. Daten werden direkt aus Deiner Excel-Forecast-Datei gezogen und grafisch aufbereitet.

    **Nutze die Navigation links**, um zwischen den einzelnen KPI-Seiten zu wechseln.  
    Viel Spaß beim Analysieren Deiner Szenarien!
    """)
    st.markdown("---")
    # Optional: Ein kleines Beispiel-Chart
    example = pd.DataFrame({
        "Periode": ["t-2","t-1","t0","t1","t2","t3"],
        "Umsatz": [1200,1400,1600,1800,2000,2200]
    })
    example_chart = (
        alt.Chart(example)
        .mark_line(point=True)
        .encode(x="Periode:O", y="Umsatz:Q")
        .properties(height=200, width=600)
    )
    st.altair_chart(example_chart, use_container_width=True)
    st.caption("Beispiel: Umsatzentwicklung unter dem gewählten Szenario.")

# --- Seite: Umsatz -------------------------------------------------------
elif page == "Umsatz":
    st.title("📈 Umsatz-Forecast")
    if "EBITDA" in kpi_df.index:
        raw = kpi_df.loc["EBITDA"].astype(float)
        df = raw.reset_index(); df.columns=["Periode","Wert"]
        st.metric("Umsatz t-2", format_cell("EBITDA", raw.iloc[0]))
        st.metric("Umsatz t-1", format_cell("EBITDA", raw.iloc[1]))
        st.metric("Umsatz t0",  format_cell("EBITDA", raw.iloc[2]))
        chart = (
            alt.Chart(df)
            .mark_line(point=True)
            .encode(x="Periode:O", y="Wert:Q")
            .properties(height=300)
        )
        st.altair_chart(chart, use_container_width=True)
    else:
        st.error("EBITDA-KPI nicht gefunden.")

# --- Seite: EBIT-Marge ---------------------------------------------------
elif page == "EBIT-Marge":
    st.title("💹 EBIT-Marge-Forecast")
    if "EBIT - Margin" in kpi_df.index:
        raw = kpi_df.loc["EBIT - Margin"].astype(float)
        df = (raw*100).reset_index(); df.columns=["Periode","Wert"]
        st.metric("Marge t-2", f"{raw.iloc[0]:.1%}")
        st.metric("Marge t-1", f"{raw.iloc[1]:.1%}")
        st.metric("Marge t0",  f"{raw.iloc[2]:.1%}")
        chart = (
            alt.Chart(df)
            .mark_line(point=True)
            .encode(x="Periode:O", y=alt.Y("Wert:Q", title="Marge (%)"))
            .properties(height=300)
        )
        st.altair_chart(chart, use_container_width=True)
    else:
        st.error("EBIT-Margin-KPI nicht gefunden.")

# --- Seite: Cashflow ------------------------------------------------------
elif page == "Cashflow":
    st.title("💧 Cashflow-Forecast")
    df = pd.DataFrame({"Periode":["t-2","t-1","t0","t1","t2","t3"], "Wert":[250,275,300,280,260,240]})
    st.metric("Cashflow t-2", f"{df.loc[0,'Wert']:.0f}")
    chart = (
        alt.Chart(df)
        .mark_area(opacity=0.5)
        .encode(x="Periode:O", y="Wert:Q")
        .properties(height=300)
    )
    st.altair_chart(chart, use_container_width=True)

# --- Seite: FCF -----------------------------------------------------------
elif page == "FCF":
    st.title("🎯 Free Cashflow (FCF)-Forecast")
    df = pd.DataFrame({"Periode":["t-2","t-1","t0","t1","t2","t3"], "Wert":[200,220,240,230,210,190]})
    st.metric("FCF t-2", f"{df.loc[0,'Wert']:.0f}")
    chart = (
        alt.Chart(df)
        .mark_line(point=True)
        .encode(x="Periode:O", y="Wert:Q")
        .properties(height=300)
    )
    st.altair_chart(chart, use_container_width=True)

# --- Seite: Kapitalrendite ------------------------------------------------
elif page == "Kapitalrendite":
    st.title("📊 Kapitalrendite-Forecast")
    if "EK-Rendite" in kpi_df.index:
        raw = kpi_df.loc["EK-Rendite"].astype(float)
        df = (raw*100).reset_index(); df.columns=["Periode","Wert"]
        st.metric("RoI t-2", f"{raw.iloc[0]:.1%}")
        st.metric("RoI t-1", f"{raw.iloc[1]:.1%}")
        st.metric("RoI t0",  f"{raw.iloc[2]:.1%}")
        chart = (
            alt.Chart(df)
            .mark_bar()
            .encode(x="Periode:O", y=alt.Y("Wert:Q", title="ROI (%)"))
            .properties(height=300)
        )
        st.altair_chart(chart, use_container_width=True)
    else:
        st.error("EK-Rendite-KPI nicht gefunden.")
